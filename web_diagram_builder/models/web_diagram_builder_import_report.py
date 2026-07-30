# Copyright 2024 TechnoLibre
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import base64
import csv
import io

from odoo import _, fields, models


class WebDiagramBuilderImportReport(models.TransientModel):
    _name = "web.diagram.builder.import.report"
    _description = "Import Report Download"

    result_id = fields.Many2one(
        "web.diagram.builder.import.result",
        required=True,
    )
    file_format = fields.Selection(
        [("csv", "CSV"), ("xlsx", "Excel (.xlsx)")],
        string="Format",
        default="csv",
        required=True,
    )

    def action_download(self):
        self.ensure_one()
        lines = self.result_id.line_ids
        rows = [["File", "Status", "Diagram Created", "Error"]]
        for line in lines:
            rows.append([
                line.file_name,
                line.status,
                line.builder_id.name if line.builder_id else "",
                line.error_message or "",
            ])

        if self.file_format == "csv":
            buf = io.StringIO()
            writer = csv.writer(buf)
            for row in rows:
                writer.writerow(row)
            file_bytes = buf.getvalue().encode("utf-8")
            mimetype = "text/csv"
            filename = "import_report.csv"
        else:
            try:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
            except ImportError:
                raise models.UserError(
                    _("openpyxl is required to export Excel files.")
                )
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Import Report"
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["D"].width = 80
            green = PatternFill("solid", fgColor="C6EFCE")
            red = PatternFill("solid", fgColor="FFC7CE")
            wrap = Alignment(wrap_text=True, vertical="top")
            for i, row in enumerate(rows):
                ws.append(row)
                if i == 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                else:
                    fill = green if row[1] == "success" else red
                    for cell in ws[i + 1]:
                        cell.fill = fill
                for cell in ws[i + 1]:
                    cell.alignment = wrap
            buf = io.BytesIO()
            wb.save(buf)
            file_bytes = buf.getvalue()
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = "import_report.xlsx"

        attachment = self.env["ir.attachment"].create({
            "name": filename,
            "datas": base64.b64encode(file_bytes),
            "mimetype": mimetype,
        })
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
